import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle
import os

def check_month_cancelled(elementos_por_mes, resultado):
    # Verifica si hay algún mes con 3 o más elementos
    client_failled = False
    for mes in elementos_por_mes.values():
        if mes["cantidad"] >= 3:
            datos_a_llenar = [resultado["order_vendor_dbname"], str(mes['month']), str(mes['cantidad']),  str(mes['shipping_list']).replace('[','').replace(']','')]
            llenar_celdas(datos_a_llenar)

            notification_msg = f"Hay {mes['cantidad']} elementos en el mes {mes['month']} del campo 'shipping_date'. Los productos son:{str(mes['shipping_list']).replace('[','').replace(']','')}"
            print(notification_msg)
            with open ("UsuariosRechazados.txt", "a") as usuario:
                usuario.write(resultado["order_vendor_dbname"] +": " +notification_msg+"\n")
            client_failled = True
            # SI TENDRIA CADA EMAIL DEL USUARIO AQUI PODRIA ENVIAR CADA NOTIFICACION
            # enviar_correo("inakielhaiek@gmail.com","Notificacion", notification_msg)
    return client_failled
        


def enviar_correo(destinatario, asunto, mensaje, archivo_adjunto):
    try:
        # PONER TU USUARIO 
        remitente = ''
        # PONER CONTRASEÑA 
        contraseña = ""

        # Crear objeto de mensaje
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = destinatario
        msg['Subject'] = asunto

        # Agregar cuerpo del mensaje
        msg.attach(MIMEText(mensaje, 'plain'))
        if archivo_adjunto != "no attachment":
            # Abrir archivo adjunto en modo binario
            with open(archivo_adjunto, 'rb') as f:
                # Crear objeto MIME base
                adjunto_MIME = MIMEBase('application', 'octet-stream')
                adjunto_MIME.set_payload(f.read())

            # Codificar el archivo adjunto en caracteres ASCII para enviarlo por correo electrónico
            encoders.encode_base64(adjunto_MIME)

            # Agregar encabezado de archivo adjunto
            adjunto_MIME.add_header('Content-Disposition', f'attachment; filename= {archivo_adjunto}')

            # Agregar archivo adjunto al mensaje de correo electrónico
            msg.attach(adjunto_MIME)

        # Crear objeto de servidor SMTP
        server = smtplib.SMTP('smtp.gmail.com', 587)

        # Iniciar sesión en el servidor SMTP
        server.starttls()
        server.login(remitente, contraseña)

        # Enviar correo electrónico
        text = msg.as_string()
        server.sendmail(remitente, destinatario, text)

        # Cerrar conexión con el servidor SMTP
        server.quit()
    except smtplib.SMTPException as e:
        print(f"Error al enviar el correo: {e}")



def llenar_celdas(datos):
    # Cargar el libro de trabajo existente
    libro = openpyxl.load_workbook('UsuariosRechazados.xlsx')
    hoja = libro.active

    # Obtener la fila actual para llenar
    fila = hoja.max_row + 1

    # Llenar las celdas con datos y colores
    for columna, (dato, color) in enumerate(zip(datos, [PatternFill(start_color="ccd5ae", end_color="ccd5ae", fill_type="solid"),
                                                      PatternFill(start_color="f4a261", end_color="f4a261", fill_type="solid"),
                                                      PatternFill(start_color="90e0ef", end_color="90e0ef", fill_type="solid"),
                                                      PatternFill(start_color="e7c6ff", end_color="e7c6ff", fill_type="solid")]), start=1):
        celda = hoja.cell(row=fila, column=columna, value=dato)
        celda.fill = color

    # Guardar el libro de trabajo actualizado
    libro.save('UsuariosRechazados.xlsx')

def crear_excel_con_formato(nombre_archivo='UsuariosRechazados.xlsx'):
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    libro = openpyxl.Workbook()
    hoja = libro.active

    # Definir los encabezados y colores
    encabezados = ['Usuario', 'Mes', 'Cantidad', 'IDs de Compra']
    colores = [PatternFill(start_color="dad7cd", end_color="dad7cd", fill_type="solid"),  # Azul
               PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Rojo
               PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Verde
               PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")]  # Violeta

    # Agregar encabezados a la hoja de cálculo y aplicar colores
    for columna, (encabezado, color) in enumerate(zip(encabezados, colores), start=1):
        celda = hoja.cell(row=1, column=columna, value=encabezado)
        celda.fill = color

    # Establecer un estilo para el encabezado
    estilo_encabezado = NamedStyle(name='encabezado')
    estilo_encabezado.font = openpyxl.styles.Font(bold=True)
    estilo_encabezado.fill = PatternFill(start_color="dad7cd", end_color="dad7cd", fill_type="solid")

    # Aplicar el estilo al encabezado (primera fila)
    for celda in hoja['1']:
        celda.style = estilo_encabezado

    # Fijar los encabezados (frozen panes)
    hoja.freeze_panes = 'A2'

    # Guardar el libro de trabajo
    libro.save(nombre_archivo)



def borrar_archivo_excel(nombre_archivo):
    if os.path.exists(nombre_archivo):
        os.remove(nombre_archivo)
        print(f"Archivo {nombre_archivo} borrado exitosamente.")
    else:
        print(f"El archivo {nombre_archivo} no existe.")